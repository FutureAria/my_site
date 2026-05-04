import html
import json
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment


ROOT = Path(__file__).resolve().parents[1]
PORTFOLIO_PATH = ROOT / "data" / "portfolio.json"
TEMPLATE_DIR = Path("/Users/juyoung/Desktop/MajorLink")
OUTPUT_DIR = ROOT / "public" / "uploads" / "project-docs"


DOC_TITLES = {
    "mvp": "MVP 기능 정의서",
    "schedule": "개발 일정표",
    "roles": "팀 역할 분담표",
    "business": "수익 모델 & 비용 계획",
    "effort": "기능별 팀 투입 계획",
    "team": "학과별 팀 구성",
    "meeting": "기획 회의록",
    "ui": "UI 기획 체크리스트",
    "qa": "Q&A 예상 질문",
    "budget": "예산 & 비용 계획",
}


PROJECT_NOTES = {
    "수강신청": {
        "problem": "동시 수강신청에서 정원 초과와 신청 데이터 불일치가 발생할 수 있음",
        "target": "수강신청 흐름을 학습하거나 PHP/MySQL 기반 신청 시스템을 검토하는 사용자",
        "market": "교육 행정·수강신청 도메인에서 정원, 재고, 좌석처럼 수량이 중요한 기능으로 확장 가능",
        "core": "장바구니에서 실제 신청까지 이어지는 트랜잭션 기반 처리 흐름",
        "backend": "PHP 서버 로직, MySQL 스키마, 트랜잭션 및 행 잠금",
        "frontend": "강의 목록, 장바구니, 신청 내역 화면",
        "risk": "동시 요청 테스트 범위가 작으면 실제 트래픽 상황의 병목을 모두 확인하기 어려움",
    },
    "물품": {
        "problem": "등록·수정·삭제 후 화면 상태와 서버 데이터가 어긋날 수 있음",
        "target": "소규모 물품 목록을 관리하거나 CRUD 기본 흐름을 확인하려는 사용자",
        "market": "재고·비품·대여 목록 관리처럼 기본 CRUD가 필요한 내부 관리 시스템으로 확장 가능",
        "core": "API 응답 기준으로 목록 상태를 다시 갱신하는 CRUD 관리 흐름",
        "backend": "REST API 응답 구조와 데이터 변경 흐름",
        "frontend": "React 컴포넌트 상태 관리, 목록/폼/상세 모달 UI",
        "risk": "데이터가 많아질 경우 검색, 페이지네이션, 권한 관리가 추가로 필요함",
    },
    "상권": {
        "problem": "공공 API 데이터의 결측치·이상치 때문에 분석 결과의 신뢰도가 흔들릴 수 있음",
        "target": "창업 후보 지역을 비교하거나 상권 데이터를 빠르게 해석하려는 예비 창업자",
        "market": "지역 상권 분석, 창업 입지 추천, 공공 데이터 시각화 서비스로 확장 가능",
        "core": "공공 데이터 수집·정제·분석·추천 결과 제공 흐름",
        "backend": "Spring Boot API, 데이터 정제, MySQL 저장 구조",
        "frontend": "지도 기반 상권 선택, 점수 카드, 차트 UI",
        "risk": "공공 데이터 갱신 주기와 분석 기준을 계속 관리해야 함",
    },
    "KIS": {
        "problem": "자동매매 판단의 이유, 실패 원인, 실거래 전환 위험을 사용자가 이해하기 어려움",
        "target": "모의투자 기반 AI 판단 흐름과 리스크 제어 구조를 확인하려는 사용자",
        "market": "개인 투자 보조 도구, 모의투자 검증 대시보드, 설명 가능한 자동화 시스템으로 확장 가능",
        "core": "시장 데이터·뉴스·모의투자 로그 기반 AI 판단과 읽기 전용 대시보드",
        "backend": "Python 데이터 수집, 스케줄링, 로그 저장, 잠금 구조",
        "frontend": "관리자/읽기모드 화면, 판단 이유 표시 UI",
        "risk": "금융 도메인은 실거래 전환 전 충분한 검증과 보수적인 잠금 구조가 필수임",
    },
    "BASE": {
        "problem": "티켓 예매, NFT 발급, QR 검표, 재판매 상태가 서로 어긋날 수 있음",
        "target": "블록체인 기반 티켓팅과 팬 자산 거래 흐름을 확인하려는 사용자",
        "market": "스포츠 티켓팅, NFT 멤버십, 공식 재판매 플랫폼으로 확장 가능",
        "core": "야구 경기 예매부터 NFT 티켓, QR 검표, 공식 재판매까지 이어지는 상태 흐름",
        "backend": "Node.js/Express, MariaDB, 인증, 티켓 상태 API",
        "frontend": "React 예매 화면, 마이티켓, 장터, 커뮤니티 UI",
        "risk": "온체인 결과와 서버 DB 상태를 보정하는 예외 처리가 계속 필요함",
    },
    "음악": {
        "problem": "AI 감정 분석 응답이 흔들리면 추천 결과와 사용자 기록도 불안정해질 수 있음",
        "target": "감정 기반 개인화 추천 경험을 원하는 사용자",
        "market": "음악 추천, 감정 기록, 개인화 큐레이션 서비스로 확장 가능",
        "core": "감정 분석 결과를 정규화해 음악 추천과 히스토리로 연결하는 흐름",
        "backend": "Express API, Gemini 연동, Firebase Admin, 추천 서비스",
        "frontend": "React 감정 입력, 추천 결과, 히스토리 UI",
        "risk": "AI 응답 품질과 사용자 데이터 보호 기준을 함께 관리해야 함",
    },
}


def find_template() -> Path:
    matches = sorted(TEMPLATE_DIR.glob("MajorLink_*.xlsx"))
    if not matches:
        raise FileNotFoundError("MajorLink 기획자료 템플릿을 찾을 수 없습니다.")
    return matches[0]


def slugify(title: str) -> str:
    mapping = [
        ("수강신청", "course-registration"),
        ("물품", "crud"),
        ("상권", "commercial-district"),
        ("KIS", "kis-ai-trader"),
        ("BASE", "base-chain"),
        ("음악", "emotion-music"),
    ]
    for key, value in mapping:
        if key in title:
            return value
    slug = re.sub(r"[^a-z0-9]+", "-", title.lower()).strip("-")
    return slug or "project"


def note_for(project: dict) -> dict:
    title = project["title"]
    for key, value in PROJECT_NOTES.items():
        if key in title:
            return value
    return {
        "problem": project.get("desc", "프로젝트 문제 정의 확인 필요"),
        "target": "포트폴리오 방문자와 프로젝트 평가자",
        "market": "포트폴리오 검토와 기술 역량 설명 자료로 활용",
        "core": project.get("desc", "핵심 기능"),
        "backend": "API, 데이터 구조, 배포 흐름",
        "frontend": "사용자 화면, 상태 관리, 반응형 UI",
        "risk": "운영 환경과 데이터 범위에 따라 추가 검증 필요",
    }


def clear_sheet(ws):
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.value = None


def set_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    cell.value = value
    cell.alignment = Alignment(
        horizontal=cell.alignment.horizontal or "left",
        vertical="center",
        wrap_text=True,
    )


def write_row(ws, row, values):
    for col, value in enumerate(values, start=1):
        set_cell(ws, row, col, value)


def project_features(project: dict) -> list[str]:
    features = project.get("detail", {}).get("features") or []
    if features:
        return features[:12]
    return project.get("techs", [])[:8]


def rewrite_mvp(ws, project, note):
    clear_sheet(ws)
    set_cell(ws, 1, 1, f"{project['title']}  |  {DOC_TITLES['mvp']}")
    set_cell(ws, 2, 1, f"{project['period']} 기준  |  MVP 범위 = {note['core']}")
    write_row(ws, 3, ["#", "기능 카테고리", "세부 기능명", "기능 설명", "담당 역할", "우선순위", "난이도", "예상 기간", "상태", "비고"])
    for idx, feature in enumerate(project_features(project), start=1):
        category = "핵심 기능" if idx <= 4 else "확장 기능"
        role = "백엔드+프론트" if idx <= 5 else "프론트/운영"
        priority = "P0 필수" if idx <= 4 else "P1 중요"
        difficulty = "중간" if idx <= 6 else "낮음"
        status = "구현 완료" if "진행 중" not in project["period"] and "개발 중" not in project["title"] else "진행 중"
        write_row(ws, 3 + idx, [f"F{idx:02d}", category, feature[:28], feature, role, priority, difficulty, "2~4일", status, note["risk"] if idx == 1 else ""])


def rewrite_schedule(ws, project, note):
    clear_sheet(ws)
    set_cell(ws, 1, 1, f"{project['title']}  |  개발 일정표 ({project['period']})")
    write_row(ws, 3, ["담당자", "분야", "역할 / 작업", "W1", "W2", "W3", "W4", "W5", "W6", "W7", "W8", "W9", "W10"])
    tasks = [
        ("전체", "기획", "문제 정의 및 MVP 범위 확정", [4]),
        ("백엔드", "서버/DB", note["backend"], [4, 5]),
        ("프론트", "UI", note["frontend"], [5, 6]),
        ("개발", "핵심 기능", note["core"], [6, 7]),
        ("개발", "연동", "프론트엔드와 API 연동 및 상태 흐름 검증", [7, 8]),
        ("전체", "테스트", "주요 시나리오 테스트와 오류 수정", [8]),
        ("운영", "배포", "GitHub 정리, 데모 배포, 포트폴리오 반영", [9]),
        ("전체", "발표", "기획자료, README, 발표 Q&A 정리", [10, 11]),
    ]
    for r, (owner, field, task, weeks) in enumerate(tasks, start=4):
        values = [owner, field, task] + ["●" if c in weeks else "" for c in range(4, 14)]
        write_row(ws, r, values)


def rewrite_roles(ws, project, note):
    clear_sheet(ws)
    set_cell(ws, 1, 1, f"{project['title']}  |  팀 역할 분담표")
    set_cell(ws, 2, 1, f"프로젝트 성격에 맞춘 역할과 산출물 정리  |  기간: {project['period']}")
    write_row(ws, 3, ["학과 / 전공", "역할", "인원", "기여도", "주요 업무 상세", "주요 산출물", "협업 도구", "비고"])
    rows = [
        ["컴퓨터공학 / 소프트웨어", "백엔드 개발", "1명", "주담당", note["backend"], "API, DB 구조, 서버 로직, 배포 기록", "GitHub, 터미널, API 테스트", "핵심"],
        ["컴퓨터공학 / 소프트웨어", "프론트엔드 개발", "1명", "주담당", note["frontend"], "화면, 컴포넌트, 반응형 UI", "GitHub, Figma, 브라우저", "핵심"],
        ["데이터 / AI", "데이터·AI 로직", "0~1명", "프로젝트별", "데이터 수집, 정제, 추천/판단 기준 설계", "분석 기준, 모델/규칙, 검증 로그", "Python, Notebook, 문서", "해당 시"],
        ["기획 / PM", "서비스 기획", "1명", "보조", note["problem"], "기능 정의서, 일정표, Q&A", "Excel, Notion", "발표"],
        ["디자인 / UX", "UI/UX 설계", "0~1명", "보조", "사용자 흐름과 화면 구조 정리", "화면 체크리스트, 와이어프레임", "Figma", "해당 시"],
    ]
    for r, row in enumerate(rows, start=4):
        write_row(ws, r, row)


def rewrite_business(ws, project, note):
    clear_sheet(ws)
    set_cell(ws, 1, 1, f"{project['title']}  |  수익 모델 & 비용 계획")
    set_cell(ws, 2, 1, "포트폴리오 프로젝트 기준의 확장 가능 모델입니다. 실제 매출 수치는 추정이며 검증이 필요합니다.")
    write_row(ws, 3, ["수익 모델", "단계", "대상", "가격대", "주요 제공 기능", "예상 전환율", "비고"])
    rows = [
        ["포트폴리오 데모", "현재", "채용 담당자/면접관", "무료", "데모, GitHub, 기획자료 제공", "해당 없음", "취업 포트폴리오 목적"],
        ["운영형 SaaS", "확장", note["target"], "추정 필요", note["core"], "확인 필요", note["market"]],
        ["B2B/교육용 패키지", "확장", "학교/동아리/소규모 조직", "협의", "관리 기능, 문서화, 배포 지원", "확인 필요", "프로젝트 성숙도에 따라 판단"],
    ]
    for r, row in enumerate(rows, start=4):
        write_row(ws, r, row)
    set_cell(ws, 9, 1, "MVP 개발 예상 비용")
    write_row(ws, 10, ["비용 항목", "분류", "월 비용 (예상)", "프로젝트 기간 합계", "무료 대안", "비고", ""])
    cost_rows = [
        ["GitHub", "코드 관리", "무료", "무료", "Free Plan", "공개 저장소 기준", ""],
        ["Oracle Cloud/Dothome", "배포", "무료~소액", "확인 필요", "무료 티어", "운영 환경별 차이", ""],
        ["AI/API 사용료", "외부 API", "사용량 기반", "확인 필요", "무료 할당량", "KIS/상권/음악 프로젝트 중심", ""],
        ["도메인/DNS", "운영", "무료", "무료", "DuckDNS", "현재 포트폴리오 운영 기준", ""],
    ]
    for r, row in enumerate(cost_rows, start=11):
        write_row(ws, r, row)


def rewrite_effort(ws, project, note):
    clear_sheet(ws)
    set_cell(ws, 1, 1, f"{project['title']}  |  기능별 팀 투입 계획")
    set_cell(ws, 2, 1, "역할: ● 주담당  ○ 보조  - 없음  |  공수는 포트폴리오 정리용 추정치")
    write_row(ws, 4, ["기능 ID", "카테고리", "기능명", "백엔드 담당", "공수(h)", "프론트 담당", "공수(h)", "UI/UX", "공수(h)", "기획/PM", "공수(h)", "문서화", "공수(h)", "합계(h)", "작업 내용 (백엔드)", "작업 내용 (프론트/기획)"])
    for idx, feature in enumerate(project_features(project), start=1):
        backend = "●" if idx % 3 != 0 else "○"
        frontend = "●" if idx % 3 != 1 else "○"
        total = 12 + (idx % 4) * 4
        write_row(ws, 4 + idx, [f"F{idx:02d}", "핵심" if idx <= 4 else "확장", feature[:28], backend, total // 3, frontend, total // 3, "○", 2, "○", 2, "○", 2, total, note["backend"], note["frontend"]])


def rewrite_team(ws, project, note):
    clear_sheet(ws)
    set_cell(ws, 1, 1, f"{project['title']} 학과별 팀 구성")
    set_cell(ws, 2, 1, "프로젝트에 필요한 전공 계열과 역할을 정리합니다.")
    write_row(ws, 4, ["구분", "학과", "", "", "", "핵심 담당 업무", "", "", "", "", "", "", "필요 인원", "", "비고", ""])
    rows = [
        ["필수", "컴퓨터공학 / 소프트웨어", "", "", "", f"{note['backend']} / {note['frontend']}", "", "", "", "", "", "", "1~2명", "", "핵심", ""],
        ["권장", "데이터사이언스 / 통계", "", "", "", "데이터 분석, 추천 기준, 검증 지표 설계", "", "", "", "", "", "", "0~1명", "", "데이터 프로젝트 중심", ""],
        ["권장", "디자인 / UX", "", "", "", "사용자 흐름, 화면 설계, 시각화 개선", "", "", "", "", "", "", "0~1명", "", "UI 개선", ""],
        ["선택", "경영 / 기획", "", "", "", "시장성, 수익 모델, 발표 스토리 정리", "", "", "", "", "", "", "0~1명", "", "발표 자료", ""],
    ]
    for r, row in enumerate(rows, start=5):
        write_row(ws, r, row)


def rewrite_meeting(ws, project, note):
    clear_sheet(ws)
    set_cell(ws, 1, 1, f"{project['title']} 기획 회의록")
    set_cell(ws, 2, 1, "프로젝트 진행 단계별 결정 사항을 기록합니다.")
    write_row(ws, 3, ["No", "날짜", "회의 유형", "참석 분야", "", "회의 주제", "", "결정 사항", "", "담당", "다음 액션", "완료 여부"])
    rows = [
        [1, project["period"].split("~")[0].strip(), "킥오프", "전체", "", "문제 정의와 MVP 범위 확정", "", note["problem"], "", "전체", "핵심 기능 목록 작성", "완료"],
        [2, "개발 초반", "기술 설계", "백엔드/프론트", "", "데이터 구조와 화면 흐름 결정", "", note["core"], "", "개발", "API와 UI 연결", "완료" if "진행 중" not in project["period"] else "진행중"],
        [3, "개발 중반", "중간 점검", "전체", "", "기능 구현 현황과 문제 해결", "", project.get("detail", {}).get("challenges", "").split("\n\n")[0][:120], "", "전체", "트러블슈팅 정리", "완료" if "진행 중" not in project["period"] else "진행중"],
        [4, "정리 단계", "발표/문서", "전체", "", "GitHub, 데모, 포트폴리오 문서화", "", "README와 기획자료를 공개 가능한 범위로 정리", "", "전체", "포트폴리오 반영", "진행중"],
    ]
    for r, row in enumerate(rows, start=4):
        write_row(ws, r, row)


def rewrite_ui(ws, project, note):
    clear_sheet(ws)
    set_cell(ws, 1, 1, f"{project['title']} UI 기획 체크리스트")
    set_cell(ws, 2, 1, "화면별 UI 구성과 개발 연계 포인트를 정리합니다.")
    write_row(ws, 3, ["No", "화면 구분", "화면명", "", "담당 분야", "", "주요 컴포넌트 / UI 요소", "", "", "디자인 요구사항", "", "상태", "개발 연계", "완료 여부"])
    screens = ["메인/대시보드", "목록/검색", "상세/결과", "입력/관리", "오류/빈 상태", "모바일 반응형"]
    for idx, screen in enumerate(screens, start=1):
        write_row(ws, 3 + idx, [idx, "핵심 화면", screen, "", "프론트/디자인", "", f"{project['title']}의 {screen} 화면", "", "", "정보가 겹치지 않고 핵심 행동이 먼저 보이도록 구성", "", "완료" if idx <= 4 else "점검", note["frontend"], "완료" if "개발 중" not in project["title"] else "진행중"])


def rewrite_qa(ws, project, note):
    clear_sheet(ws)
    set_cell(ws, 1, 1, f"{project['title']} Q&A 예상 질문 & 모범 답변")
    set_cell(ws, 2, 1, "발표 또는 면접에서 받을 수 있는 질문과 답변을 정리합니다.")
    write_row(ws, 3, ["No", "카테고리", "난이도", "예상 질문", "", "", "모범 답변", "", "", "", ""])
    questions = [
        ("서비스/기획", "★★★", "이 프로젝트가 해결하려는 핵심 문제는 무엇인가요?", note["problem"]),
        ("기술/개발", "★★★", "백엔드 관점에서 가장 중요하게 설계한 부분은 무엇인가요?", note["backend"]),
        ("기술/개발", "★★", "프론트엔드 관점에서 사용성을 어떻게 고려했나요?", note["frontend"]),
        ("데이터/운영", "★★★", "데이터나 상태가 어긋나는 문제는 어떻게 줄였나요?", project.get("detail", {}).get("challenges", "").split("\n\n")[1].replace("해결:", "").strip()[:240] if "해결:" in project.get("detail", {}).get("challenges", "") else note["core"]),
        ("보안/운영", "★★", "공개 저장소에 올릴 때 어떤 점을 조심했나요?", "API Key, DB 비밀번호, 운영 환경 파일, 실제 사용자 데이터는 제외하고 README와 예시 설정 중심으로 공개했습니다."),
        ("확장성", "★★", "앞으로 개선한다면 무엇을 먼저 하겠나요?", note["risk"]),
    ]
    for idx, q in enumerate(questions, start=1):
        write_row(ws, 3 + idx, [idx, q[0], q[1], q[2], "", "", q[3], "", "", "", ""])


def rewrite_budget(ws, project, note):
    clear_sheet(ws)
    set_cell(ws, 1, 1, f"{project['title']} 예산 & 비용 계획")
    set_cell(ws, 2, 1, "포트폴리오 데모 운영 기준의 예상 비용입니다. 실제 비용은 사용량과 배포 환경에 따라 달라질 수 있습니다.")
    write_row(ws, 4, ["No", "항목", "", "", "구분", "월 비용", "연간 비용", "무료 플랜", "비고", ""])
    rows = [
        [1, "GitHub", "", "", "코드 관리", "무료", "무료", "O", "공개 저장소", ""],
        [2, "Oracle Cloud / Dothome", "", "", "배포", "무료~소액", "확인 필요", "O", "현재 데모 운영 기준", ""],
        [3, "DuckDNS", "", "", "도메인", "무료", "무료", "O", "서브도메인 사용", ""],
        [4, "DB / 스토리지", "", "", "데이터", "무료~소액", "확인 필요", "△", "프로젝트별 데이터 크기에 따라 변동", ""],
        [5, "외부 API", "", "", "API", "사용량 기반", "확인 필요", "△", "AI/공공/금융 API 사용 시", ""],
        [6, "모니터링/백업", "", "", "운영", "무료~소액", "확인 필요", "△", "서비스 확장 시 필요", ""],
    ]
    for r, row in enumerate(rows, start=5):
        write_row(ws, r, row)


REWRITERS = [
    rewrite_mvp,
    rewrite_schedule,
    rewrite_roles,
    rewrite_business,
    rewrite_effort,
    rewrite_team,
    rewrite_meeting,
    rewrite_ui,
    rewrite_qa,
    rewrite_budget,
]


def render_html(workbook_path: Path, project: dict, xlsx_public_path: str) -> str:
    wb = load_workbook(workbook_path, data_only=True)
    sections = []
    for ws in wb.worksheets:
        rows_html = []
        for row in ws.iter_rows():
            values = [cell.value for cell in row]
            if not any(value not in (None, "") for value in values):
                continue
            cells = []
            for value in values:
                tag = "th" if row[0].row <= 3 else "td"
                text = "" if value is None else str(value)
                cells.append(f"<{tag}>{html.escape(text)}</{tag}>")
            rows_html.append("<tr>" + "".join(cells) + "</tr>")
        sections.append(
            f"<section><h2>{html.escape(ws.title)}</h2><div class=\"table-wrap\"><table>{''.join(rows_html)}</table></div></section>"
        )
    title = html.escape(project["title"])
    return f"""<!doctype html>
<html lang=\"ko\">
<head>
  <meta charset=\"utf-8\" />
  <meta name=\"viewport\" content=\"width=device-width, initial-scale=1\" />
  <title>{title} 기획자료</title>
  <style>
    :root {{ color-scheme: dark; --bg:#050812; --panel:#0f172a; --line:#273449; --text:#e5edf8; --muted:#9aa8bc; --accent:#5eead4; }}
    * {{ box-sizing: border-box; }}
    body {{ margin:0; font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif; background:var(--bg); color:var(--text); line-height:1.6; }}
    header {{ position:sticky; top:0; z-index:1; padding:18px 20px; border-bottom:1px solid var(--line); background:rgba(5,8,18,.92); backdrop-filter:blur(12px); }}
    main {{ max-width:1180px; margin:0 auto; padding:24px 16px 56px; }}
    h1 {{ margin:0; font-size:20px; letter-spacing:0; }}
    .meta {{ margin-top:4px; color:var(--muted); font-size:13px; }}
    .actions {{ margin-top:12px; display:flex; gap:8px; flex-wrap:wrap; }}
    a.button {{ display:inline-flex; align-items:center; gap:6px; padding:8px 12px; border-radius:8px; border:1px solid var(--line); color:var(--text); text-decoration:none; font-size:13px; background:#111827; }}
    section {{ margin:0 0 24px; }}
    h2 {{ margin:0 0 10px; color:var(--accent); font-size:15px; }}
    .table-wrap {{ overflow:auto; border:1px solid var(--line); border-radius:8px; background:var(--panel); }}
    table {{ width:100%; border-collapse:collapse; min-width:900px; }}
    th,td {{ border:1px solid rgba(148,163,184,.18); padding:9px 10px; vertical-align:top; font-size:13px; white-space:pre-wrap; }}
    th {{ background:#162033; color:#dbeafe; font-weight:700; }}
    td {{ color:#d3dce8; }}
  </style>
</head>
<body>
  <header>
    <h1>{title} 기획자료</h1>
    <div class=\"meta\">포트폴리오 공개용 문서 · 원본 엑셀 다운로드 가능</div>
    <div class=\"actions\"><a class=\"button\" href=\"{html.escape(xlsx_public_path)}\" download>엑셀 다운로드</a></div>
  </header>
  <main>{''.join(sections)}</main>
</body>
</html>
"""


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    template = find_template()
    data = json.loads(PORTFOLIO_PATH.read_text(encoding="utf-8"))

    for project in data["projects"]:
        slug = slugify(project["title"])
        xlsx_path = OUTPUT_DIR / f"{slug}-planning.xlsx"
        html_path = OUTPUT_DIR / f"{slug}-planning.html"
        wb = load_workbook(template)
        note = note_for(project)
        for ws, rewrite in zip(wb.worksheets, REWRITERS):
            rewrite(ws, project, note)
        wb.save(xlsx_path)
        preview = render_html(xlsx_path, project, f"/uploads/project-docs/{xlsx_path.name}")
        html_path.write_text(preview, encoding="utf-8")
        project["documents"] = [
            {
                "title": "프로젝트 기획자료",
                "description": "MVP 기능 정의서, 개발 일정표, 역할 분담표, 비용 계획, UI 체크리스트, Q&A를 프로젝트 상황에 맞게 정리한 공개용 엑셀 문서입니다.",
                "file": f"/uploads/project-docs/{xlsx_path.name}",
                "preview": f"/uploads/project-docs/{html_path.name}",
                "type": "xlsx",
            }
        ]

    PORTFOLIO_PATH.write_text(
        json.dumps(data, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
